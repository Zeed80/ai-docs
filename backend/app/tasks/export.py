"""Export Celery tasks — Excel and 1C payload generation."""

import uuid
from datetime import datetime, timezone
from io import BytesIO

import structlog
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.db.models import ExportJob, Invoice, InvoiceLine, Party
from app.tasks.celery_app import celery_app

logger = structlog.get_logger()


def _get_sync_session() -> Session:
    from sqlalchemy import create_engine
    engine = create_engine(settings.database_url_sync, pool_pre_ping=True)
    return Session(engine)


def _run_async(coro):
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as pool:
                return pool.submit(asyncio.run, coro).result()
        return loop.run_until_complete(coro)
    except RuntimeError:
        return asyncio.run(coro)


@celery_app.task(name="app.tasks.export.generate_excel_export", bind=True, max_retries=2)
def generate_excel_export(self, job_id: str) -> dict:
    """Generate Excel file for an invoice export job."""
    logger.info("excel_export_start", job_id=job_id)

    with _get_sync_session() as db:
        job = db.get(ExportJob, uuid.UUID(job_id))
        if not job:
            return {"error": "ExportJob not found"}

        job.status = "generating"
        db.commit()

        try:
            invoice = db.execute(
                select(Invoice)
                .where(Invoice.id == job.entity_id)
                .options(
                    joinedload(Invoice.lines),
                    joinedload(Invoice.supplier),
                )
            ).unique().scalar_one_or_none()

            if not invoice:
                job.status = "failed"
                job.error = "Invoice not found"
                db.commit()
                return {"error": "Invoice not found"}

            buf = _build_excel(invoice)

            # Save to storage
            filename = f"invoice_{invoice.invoice_number or job.entity_id}_export.xlsx"
            storage_path = f"exports/{job_id}/{filename}"

            from app.storage import upload_file
            upload_file(buf.getvalue(), storage_path, content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ))

            job.storage_path = storage_path
            job.status = "ready"
            job.ready_at = datetime.now(timezone.utc)
            db.commit()

            logger.info("excel_export_done", job_id=job_id, path=storage_path)
            return {"job_id": job_id, "status": "ready", "path": storage_path}

        except Exception as e:
            job.status = "failed"
            job.error = str(e)
            db.commit()
            logger.error("excel_export_error", job_id=job_id, error=str(e))
            self.retry(countdown=30, exc=e)
            return {"error": str(e)}


@celery_app.task(name="app.tasks.export.generate_1c_export", bind=True, max_retries=2)
def generate_1c_export(self, job_id: str) -> dict:
    """Generate 1C XML payload for an invoice export job."""
    logger.info("1c_export_start", job_id=job_id)

    with _get_sync_session() as db:
        job = db.get(ExportJob, uuid.UUID(job_id))
        if not job:
            return {"error": "ExportJob not found"}

        job.status = "generating"
        db.commit()

        try:
            invoice = db.execute(
                select(Invoice)
                .where(Invoice.id == job.entity_id)
                .options(
                    joinedload(Invoice.lines),
                    joinedload(Invoice.supplier),
                )
            ).unique().scalar_one_or_none()

            if not invoice:
                job.status = "failed"
                job.error = "Invoice not found"
                db.commit()
                return {"error": "Invoice not found"}

            xml_content = _build_1c_xml(invoice)
            filename = f"invoice_{invoice.invoice_number or job.entity_id}_1c.xml"
            storage_path = f"exports/{job_id}/{filename}"

            from app.storage import upload_file
            upload_file(xml_content.encode("utf-8"), storage_path, content_type="application/xml")

            job.storage_path = storage_path
            job.status = "ready"
            job.ready_at = datetime.now(timezone.utc)
            db.commit()

            logger.info("1c_export_done", job_id=job_id, path=storage_path)
            return {"job_id": job_id, "status": "ready", "path": storage_path}

        except Exception as e:
            job.status = "failed"
            job.error = str(e)
            db.commit()
            logger.error("1c_export_error", job_id=job_id, error=str(e))
            self.retry(countdown=30, exc=e)
            return {"error": str(e)}


def _build_excel(invoice: Invoice) -> BytesIO:
    """Build Excel workbook for an invoice."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Счёт"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    # Header info
    supplier_name = invoice.supplier.name if invoice.supplier else "—"
    ws["A1"] = "Счёт №"
    ws["B1"] = invoice.invoice_number or "—"
    ws["A2"] = "Поставщик"
    ws["B2"] = supplier_name
    ws["A3"] = "Дата счёта"
    ws["B3"] = invoice.invoice_date.strftime("%d.%m.%Y") if invoice.invoice_date else "—"
    ws["A4"] = "Срок оплаты"
    ws["B4"] = invoice.due_date.strftime("%d.%m.%Y") if invoice.due_date else "—"
    ws["A5"] = "Валюта"
    ws["B5"] = invoice.currency

    # Column headers
    headers = ["№", "Наименование", "Кол-во", "Ед.изм", "Цена", "Сумма", "НДС %", "НДС сумма"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Lines
    for row_idx, line in enumerate(invoice.lines, 8):
        ws.cell(row=row_idx, column=1, value=line.line_number)
        ws.cell(row=row_idx, column=2, value=line.description)
        ws.cell(row=row_idx, column=3, value=line.quantity)
        ws.cell(row=row_idx, column=4, value=line.unit)
        ws.cell(row=row_idx, column=5, value=line.unit_price)
        ws.cell(row=row_idx, column=6, value=line.amount)
        ws.cell(row=row_idx, column=7, value=line.tax_rate)
        ws.cell(row=row_idx, column=8, value=line.tax_amount)

    # Totals
    last_row = 8 + len(invoice.lines)
    ws.cell(row=last_row, column=5, value="Итого:").font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=invoice.subtotal).font = Font(bold=True)
    ws.cell(row=last_row + 1, column=5, value="НДС:").font = Font(bold=True)
    ws.cell(row=last_row + 1, column=6, value=invoice.tax_amount).font = Font(bold=True)
    ws.cell(row=last_row + 2, column=5, value="Всего к оплате:").font = Font(bold=True)
    ws.cell(row=last_row + 2, column=6, value=invoice.total_amount).font = Font(bold=True)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_1c_xml(invoice: Invoice) -> str:
    """Build a CommerceML 2.08-compatible XML document for 1C import.

    Structure: КоммерческаяИнформация / Документ / ТаблицаЧасти / Контрагент
    following the standard used by 1C:Бухгалтерия exchange format.
    """
    from datetime import timezone as _tz

    supplier = invoice.supplier
    s_name = _xml_escape(supplier.name if supplier else "")
    s_inn = _xml_escape(supplier.inn if supplier else "")
    s_kpp = _xml_escape(supplier.kpp if supplier else "")
    s_bank = _xml_escape(supplier.bank_name if supplier else "")
    s_bik = _xml_escape(supplier.bank_bik if supplier else "")
    s_account = _xml_escape(supplier.bank_account if supplier else "")
    s_corr = _xml_escape(supplier.corr_account if supplier else "")

    inv_number = _xml_escape(invoice.invoice_number or "")
    inv_date = invoice.invoice_date.strftime("%Y%m%d") if invoice.invoice_date else ""
    due_date = invoice.due_date.strftime("%Y%m%d") if invoice.due_date else ""
    currency = _xml_escape(invoice.currency or "RUB")
    created_at = (
        invoice.created_at.astimezone(_tz.utc).strftime("%Y%m%dT%H%M%S")
        if invoice.created_at else ""
    )

    # TabularSection rows — one <СтрокаТаблицыЧасти> per line item
    rows_xml = ""
    for idx, line in enumerate(invoice.lines, start=1):
        tax_rate_str = f"НДС{int((line.tax_rate or 0) * 100)}" if line.tax_rate else "БезНДС"
        rows_xml += f"""\
        <СтрокаТаблицыЧасти>
          <НомерСтроки>{line.line_number or idx}</НомерСтроки>
          <Номенклатура>
            <Наименование>{_xml_escape(line.description or '')}</Наименование>
            <Артикул>{_xml_escape(str(line.canonical_item_id or ''))}</Артикул>
          </Номенклатура>
          <ЕдиницаИзмерения>
            <НаименованиеПолное>{_xml_escape(line.unit or 'шт')}</НаименованиеПолное>
          </ЕдиницаИзмерения>
          <Количество>{line.quantity or 0}</Количество>
          <Цена>{_fmt_amount(line.unit_price)}</Цена>
          <Сумма>{_fmt_amount(line.amount)}</Сумма>
          <СтавкаНДС>{tax_rate_str}</СтавкаНДС>
          <СуммаНДС>{_fmt_amount(line.tax_amount)}</СуммаНДС>
          <Всего>{_fmt_amount((line.amount or 0) + (line.tax_amount or 0))}</Всего>
        </СтрокаТаблицыЧасти>
"""

    return f"""\
<?xml version="1.0" encoding="UTF-8"?>
<КоммерческаяИнформация
    xmlns="urn:1C.ru:commerceml_2"
    xmlns:xs="http://www.w3.org/2001/XMLSchema"
    ВерсияСхемы="2.08"
    ДатаФормирования="{created_at}">
  <Документ>
    <Ид>{invoice.id}</Ид>
    <Номер>{inv_number}</Номер>
    <Дата>{inv_date}</Дата>
    <СрокПлатежа>{due_date}</СрокПлатежа>
    <ХозяйственнаяОперация>СчетНаОплатуПоставщика</ХозяйственнаяОперация>
    <Роль>Продавец</Роль>
    <Валюта>{currency}</Валюта>
    <Сумма>{_fmt_amount(invoice.total_amount)}</Сумма>
    <СуммаНДС>{_fmt_amount(invoice.tax_amount)}</СуммаНДС>
    <Контрагенты>
      <Контрагент>
        <Ид>{supplier.id if supplier else ''}</Ид>
        <Наименование>{s_name}</Наименование>
        <ИНН>{s_inn}</ИНН>
        <КПП>{s_kpp}</КПП>
        <Роль>Продавец</Роль>
        <БанковскиеСчета>
          <БанковскийСчет>
            <Номер>{s_account}</Номер>
            <Банк>
              <Наименование>{s_bank}</Наименование>
              <БИК>{s_bik}</БИК>
              <КоррСчет>{s_corr}</КоррСчет>
            </Банк>
          </БанковскийСчет>
        </БанковскиеСчета>
      </Контрагент>
    </Контрагенты>
    <ТаблицаЧасти>
      <НаименованиеТаблицыЧасти>Товары</НаименованиеТаблицыЧасти>
{rows_xml}    </ТаблицаЧасти>
    <ЗначенияРеквизитов>
      <ЗначениеРеквизита>
        <Наименование>СуммаДокумента</Наименование>
        <Значение>{_fmt_amount(invoice.subtotal)}</Значение>
      </ЗначениеРеквизита>
      <ЗначениеРеквизита>
        <Наименование>СуммаНДС</Наименование>
        <Значение>{_fmt_amount(invoice.tax_amount)}</Значение>
      </ЗначениеРеквизита>
      <ЗначениеРеквизита>
        <Наименование>СуммаСНДС</Наименование>
        <Значение>{_fmt_amount(invoice.total_amount)}</Значение>
      </ЗначениеРеквизита>
    </ЗначенияРеквизитов>
  </Документ>
</КоммерческаяИнформация>
"""


def _fmt_amount(value: float | None) -> str:
    """Format monetary value with 2 decimal places for 1C XML."""
    return f"{value:.2f}" if value is not None else "0.00"


def _xml_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )
