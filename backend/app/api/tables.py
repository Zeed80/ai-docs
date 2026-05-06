"""Tables & Export API — skills: table.query, table.export_excel,
table.export_1c, table.import_excel, table.apply_diff"""

import io
import uuid
from datetime import datetime

import structlog
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.audit.service import log_action
from app.auth.jwt import require_role
from app.auth.models import UserInfo, UserRole
from app.db.models import (
    Document,
    DocumentStatus,
    DocumentType,
    Invoice,
    InvoiceStatus,
    SavedView,
)
from app.db.session import get_db
from app.domain.tables import (
    Export1CRequest,
    ExportRequest,
    ImportDiffResponse,
    ImportDiffRow,
    SavedViewCreate,
    SavedViewOut,
    TableColumn,
    TableFilter,
    TableQueryRequest,
    TableQueryResponse,
    TableRow,
    TableSort,
)
from app.formatting import format_money, format_number, is_money_key, to_decimal

router = APIRouter()
logger = structlog.get_logger()


# ── Column definitions ─────────────────────────────────────────────────────

INVOICE_COLUMNS = [
    TableColumn(key="invoice_number", label="Номер счёта", data_type="string"),
    TableColumn(key="invoice_date", label="Дата", data_type="date"),
    TableColumn(key="supplier_name", label="Поставщик", data_type="string"),
    TableColumn(key="total_amount", label="Сумма", data_type="number"),
    TableColumn(key="currency", label="Валюта", data_type="string"),
    TableColumn(key="tax_amount", label="НДС", data_type="number"),
    TableColumn(key="subtotal", label="Сумма без НДС", data_type="number"),
    TableColumn(key="status", label="Статус", data_type="enum"),
    TableColumn(key="overall_confidence", label="Уверенность", data_type="number"),
    TableColumn(key="line_count", label="Позиций", data_type="number"),
    TableColumn(key="created_at", label="Создан", data_type="date"),
]

DOCUMENT_COLUMNS = [
    TableColumn(key="file_name", label="Файл", data_type="string"),
    TableColumn(key="doc_type", label="Тип", data_type="enum"),
    TableColumn(key="status", label="Статус", data_type="enum"),
    TableColumn(key="file_size", label="Размер", data_type="number"),
    TableColumn(key="page_count", label="Страниц", data_type="number"),
    TableColumn(key="source_channel", label="Источник", data_type="string"),
    TableColumn(key="created_at", label="Создан", data_type="date"),
]


# ── table.query ────────────────────────────────────────────────────────────


@router.post("/query", response_model=TableQueryResponse)
async def table_query(
    payload: TableQueryRequest,
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.query — Query table with filters, sort, pagination."""
    if payload.table == "invoices":
        return await _query_invoices(payload, db)
    elif payload.table == "documents":
        return await _query_documents(payload, db)
    else:
        raise HTTPException(400, f"Unknown table: {payload.table}")


async def _query_invoices(req: TableQueryRequest, db: AsyncSession) -> TableQueryResponse:
    query = (
        select(Invoice)
        .options(selectinload(Invoice.lines), selectinload(Invoice.supplier))
    )

    # Apply filters
    for f in req.filters:
        query = _apply_invoice_filter(query, f)

    # Search
    if req.search:
        query = query.where(
            or_(
                Invoice.invoice_number.ilike(f"%{req.search}%"),
            )
        )

    # Count
    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    # Sort
    for s in req.sort:
        col = _get_invoice_sort_column(s.column)
        if col is not None:
            query = query.order_by(col.desc() if s.direction == "desc" else col.asc())
    if not req.sort:
        query = query.order_by(Invoice.created_at.desc())

    # Paginate
    query = query.offset(req.offset).limit(req.limit)
    result = await db.execute(query)
    invoices = result.scalars().all()

    # Build rows
    columns = [c for c in INVOICE_COLUMNS if not req.columns or c.key in req.columns]
    rows = []
    for inv in invoices:
        rows.append(TableRow(
            id=str(inv.id),
            data={
                "document_id": str(inv.document_id) if inv.document_id else None,
                "invoice_number": inv.invoice_number,
                "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else None,
                "supplier_name": inv.supplier.name if inv.supplier else None,
                "total_amount": inv.total_amount,
                "currency": inv.currency,
                "tax_amount": inv.tax_amount,
                "subtotal": inv.subtotal,
                "status": inv.status.value if inv.status else None,
                "overall_confidence": inv.overall_confidence,
                "line_count": len(inv.lines),
                "created_at": inv.created_at.isoformat() if inv.created_at else None,
            },
        ))

    return TableQueryResponse(
        columns=columns, rows=rows, total=total,
        offset=req.offset, limit=req.limit,
    )


def _apply_invoice_filter(query, f: TableFilter):
    col_map = {
        "status": Invoice.status,
        "currency": Invoice.currency,
        "supplier_id": Invoice.supplier_id,
        "total_amount": Invoice.total_amount,
        "tax_amount": Invoice.tax_amount,
        "invoice_date": Invoice.invoice_date,
        "overall_confidence": Invoice.overall_confidence,
    }
    col = col_map.get(f.column)
    if col is None:
        return query

    if f.operator == "eq":
        if f.column == "status":
            try:
                return query.where(col == InvoiceStatus(f.value))
            except ValueError:
                return query
        return query.where(col == f.value)
    elif f.operator == "gt":
        return query.where(col > float(f.value))
    elif f.operator == "lt":
        return query.where(col < float(f.value))
    elif f.operator == "gte":
        return query.where(col >= float(f.value))
    elif f.operator == "lte":
        return query.where(col <= float(f.value))
    elif f.operator == "contains" and isinstance(f.value, str):
        return query.where(col.ilike(f"%{f.value}%"))
    return query


def _get_invoice_sort_column(key: str):
    mapping = {
        "invoice_number": Invoice.invoice_number,
        "invoice_date": Invoice.invoice_date,
        "total_amount": Invoice.total_amount,
        "tax_amount": Invoice.tax_amount,
        "currency": Invoice.currency,
        "status": Invoice.status,
        "overall_confidence": Invoice.overall_confidence,
        "created_at": Invoice.created_at,
    }
    return mapping.get(key)


async def _query_documents(req: TableQueryRequest, db: AsyncSession) -> TableQueryResponse:
    query = select(Document)

    for f in req.filters:
        col_map = {
            "status": Document.status,
            "doc_type": Document.doc_type,
            "source_channel": Document.source_channel,
        }
        col = col_map.get(f.column)
        if col and f.operator == "eq":
            if f.column == "status":
                try:
                    query = query.where(col == DocumentStatus(f.value))
                except ValueError:
                    pass
            elif f.column == "doc_type":
                try:
                    query = query.where(col == DocumentType(f.value))
                except ValueError:
                    pass
            else:
                query = query.where(col == f.value)

    if req.search:
        query = query.where(Document.file_name.ilike(f"%{req.search}%"))

    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    query = query.order_by(Document.created_at.desc()).offset(req.offset).limit(req.limit)
    result = await db.execute(query)
    docs = result.scalars().all()

    columns = [c for c in DOCUMENT_COLUMNS if not req.columns or c.key in req.columns]
    rows = [
        TableRow(
            id=str(d.id),
            data={
                "file_name": d.file_name,
                "doc_type": d.doc_type.value if d.doc_type else None,
                "status": d.status.value if d.status else None,
                "file_size": d.file_size,
                "page_count": d.page_count,
                "source_channel": d.source_channel,
                "created_at": d.created_at.isoformat() if d.created_at else None,
            },
        )
        for d in docs
    ]

    return TableQueryResponse(
        columns=columns, rows=rows, total=total,
        offset=req.offset, limit=req.limit,
    )


# ── table.export_excel ─────────────────────────────────────────────────────


@router.post("/export", response_class=StreamingResponse)
async def export_excel(
    payload: ExportRequest,
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.export_excel — Export table to Excel/CSV."""
    # Fetch data using table.query logic
    query_req = TableQueryRequest(
        table=payload.table,
        filters=payload.filters,
        columns=payload.columns,
        limit=500,
    )

    if payload.table == "invoices":
        data = await _query_invoices(query_req, db)
    else:
        data = await _query_documents(query_req, db)

    if payload.format == "csv":
        return _export_csv(data)

    return _export_xlsx(data, payload.table)


def _export_xlsx(data: TableQueryResponse, table_name: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = table_name.capitalize()
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Header style
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_side = Side(style="thin", color="B7C9D6")
    header_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    cell_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    wrap_top = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_top = Alignment(horizontal="right", vertical="top", wrap_text=True)
    money_format = '# ##0,00'

    # Headers
    for col_idx, col in enumerate(data.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center

    # Data rows — hidden ID column at end
    for row_idx, row in enumerate(data.rows, 2):
        for col_idx, col in enumerate(data.columns, 1):
            value = _excel_cell_value(row.data.get(col.key), col.key, col.data_type)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = wrap_top
            if is_money_key(col.key):
                cell.number_format = money_format
                cell.alignment = right_top
            elif col.data_type == "number":
                cell.number_format = '# ##0,####'
                cell.alignment = right_top
        # Hidden ID column
        ws.cell(row=row_idx, column=len(data.columns) + 1, value=row.id)

    # Hide ID column
    id_col_letter = _col_letter(len(data.columns) + 1)
    ws.column_dimensions[id_col_letter].hidden = True

    # Auto-width
    for col_idx, col in enumerate(data.columns, 1):
        max_len = len(col.label)
        for row in data.rows:
            val = str(row.data.get(col.key, "") or "")
            max_len = max(max_len, max((len(line) for line in val.splitlines()), default=0))
        width = min(max(max_len + 3, 12), 60)
        if is_money_key(col.key):
            width = max(width, 16)
        ws.column_dimensions[_col_letter(col_idx)].width = width

    for row_idx in range(2, len(data.rows) + 2):
        max_lines = 1
        for col_idx in range(1, len(data.columns) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(value, str):
                max_lines = max(max_lines, value.count("\n") + 1)
        ws.row_dimensions[row_idx].height = min(max(18, max_lines * 16), 180)

    if data.columns:
        last_col = get_column_letter(len(data.columns))
        last_row = max(len(data.rows) + 1, 1)
        display_name = f"{table_name[:20].replace('-', '_')}_export"
        table = Table(displayName=display_name, ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Footer with metadata
    footer_row = len(data.rows) + 3
    ws.cell(row=footer_row, column=1, value=f"Экспорт: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="999999", size=9)
    ws.cell(row=footer_row + 1, column=1, value=f"Всего записей: {data.total}")
    ws.cell(row=footer_row + 1, column=1).font = Font(italic=True, color="999999", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{table_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_csv(data: TableQueryResponse) -> StreamingResponse:
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\n")
    writer.writerow([c.label for c in data.columns])
    for row in data.rows:
        writer.writerow(
            [_csv_cell_value(row.data.get(c.key), c.key, c.data_type) for c in data.columns]
        )

    output = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _col_letter(idx: int) -> str:
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _excel_cell_value(value, key: str, data_type: str):
    if value is None:
        return None
    if is_money_key(key) or data_type == "number":
        number = to_decimal(value)
        if number is not None:
            return float(number)
    return value


def _csv_cell_value(value, key: str, data_type: str) -> str:
    if value is None:
        return ""
    if is_money_key(key):
        return format_money(value)
    if data_type == "number":
        return format_number(value)
    return str(value)


# ── table.export_1c (CommerceML XML) ──────────────────────────────────────


@router.post("/export-1c", response_class=StreamingResponse)
async def export_1c(
    payload: Export1CRequest,
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.export_1c — Export invoices to 1С CommerceML XML format."""
    query = (
        select(Invoice)
        .options(
            selectinload(Invoice.lines),
            selectinload(Invoice.supplier),
            selectinload(Invoice.buyer),
        )
    )

    if payload.invoice_ids:
        query = query.where(Invoice.id.in_(payload.invoice_ids))

    for f in payload.filters:
        query = _apply_invoice_filter(query, f)

    result = await db.execute(query.order_by(Invoice.created_at.desc()).limit(500))
    invoices = result.scalars().all()

    if not invoices:
        raise HTTPException(404, "No invoices found for export")

    xml_content = _build_commerceml_xml(invoices)

    await log_action(
        db,
        action="table.export_1c",
        entity_type="invoice",
        entity_id=None,
        details={"count": len(invoices)},
    )
    await db.commit()

    filename = f"invoices_1c_{datetime.now().strftime('%Y%m%d_%H%M')}.xml"
    return StreamingResponse(
        io.BytesIO(xml_content.encode("utf-8")),
        media_type="application/xml; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_commerceml_xml(invoices: list) -> str:
    """Build CommerceML 2.10 compatible XML for 1С import."""
    from lxml import etree

    root = etree.Element("КоммерческаяИнформация", attrib={
        "ВерсияСхемы": "2.10",
        "ДатаФормирования": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    })

    for inv in invoices:
        doc_el = etree.SubElement(root, "Документ")

        _add_text(doc_el, "Ид", str(inv.id))
        _add_text(doc_el, "Номер", inv.invoice_number or "")
        _add_text(doc_el, "Дата", inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "")
        _add_text(doc_el, "ХозОперация", "Покупка")
        _add_text(doc_el, "Роль", "Покупатель")
        _add_text(doc_el, "Валюта", inv.currency or "RUB")
        _add_text(doc_el, "Курс", "1")
        _add_text(doc_el, "Сумма", f"{inv.total_amount:.2f}" if inv.total_amount else "0.00")

        # Counterparty
        if inv.supplier:
            agent = etree.SubElement(doc_el, "Контрагенты")
            ctr = etree.SubElement(agent, "Контрагент")
            _add_text(ctr, "Ид", str(inv.supplier.id))
            _add_text(ctr, "Наименование", inv.supplier.name)
            _add_text(ctr, "Роль", "Продавец")
            if inv.supplier.inn:
                _add_text(ctr, "ИНН", inv.supplier.inn)
            if inv.supplier.kpp:
                _add_text(ctr, "КПП", inv.supplier.kpp)

        # Line items
        items_el = etree.SubElement(doc_el, "Товары")
        for line in inv.lines:
            item_el = etree.SubElement(items_el, "Товар")
            _add_text(item_el, "Ид", str(line.id))
            _add_text(item_el, "Наименование", line.description or f"Позиция {line.line_number}")
            _add_text(item_el, "ЕдиницаИзмерения", line.unit or "шт")
            _add_text(item_el, "Количество", f"{line.quantity:.3f}" if line.quantity else "0")
            _add_text(
                item_el,
                "ЦенаЗаЕдиницу",
                f"{line.unit_price:.2f}" if line.unit_price else "0.00",
            )
            _add_text(item_el, "Сумма", f"{line.amount:.2f}" if line.amount else "0.00")

            if line.tax_rate is not None:
                tax_el = etree.SubElement(item_el, "СтавкиНалогов")
                rate_el = etree.SubElement(tax_el, "СтавкаНалога")
                _add_text(rate_el, "Наименование", "НДС")
                _add_text(rate_el, "Ставка", f"{line.tax_rate:.0f}")

            if line.tax_amount is not None:
                taxes_el = etree.SubElement(item_el, "Налоги")
                tax_item = etree.SubElement(taxes_el, "Налог")
                _add_text(tax_item, "Наименование", "НДС")
                _add_text(tax_item, "Сумма", f"{line.tax_amount:.2f}")

    return etree.tostring(
        root,
        pretty_print=True,
        xml_declaration=True,
        encoding="UTF-8",
    ).decode("utf-8")


def _add_text(parent, tag: str, text: str):
    from lxml import etree
    el = etree.SubElement(parent, tag)
    el.text = text
    return el


# ── SavedView CRUD ─────────────────────────────────────────────────────────


@router.get("/views", response_model=list[SavedViewOut])
async def list_saved_views(
    table: str = Query("invoices"),
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.list_views — List saved table views."""
    result = await db.execute(
        select(SavedView)
        .where(SavedView.entity_type == table)
        .order_by(SavedView.created_at.desc())
    )
    views = result.scalars().all()
    return [
        SavedViewOut(
            id=v.id,
            name=v.name,
            table=v.entity_type,
            columns=v.columns,
            filters=[TableFilter(**f) for f in (v.filters or {}).get("filters", [])],
            sort=[TableSort(column=v.sort_by, direction=v.sort_order)] if v.sort_by else [],
            is_shared=v.is_shared,
            created_by=v.user_id,
            created_at=v.created_at,
        )
        for v in views
    ]


@router.post("/views", response_model=SavedViewOut)
async def create_saved_view(
    payload: SavedViewCreate,
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.create_view — Create a saved table view."""
    view = SavedView(
        name=payload.name,
        entity_type=payload.table,
        user_id="user",
        filters={"filters": [f.model_dump() for f in payload.filters]},
        columns=payload.columns,
        sort_by=payload.sort[0].column if payload.sort else None,
        sort_order=payload.sort[0].direction if payload.sort else "desc",
        is_shared=payload.is_shared,
    )
    db.add(view)
    await db.commit()
    await db.refresh(view)

    return SavedViewOut(
        id=view.id,
        name=view.name,
        table=view.entity_type,
        columns=view.columns,
        filters=payload.filters,
        sort=payload.sort,
        is_shared=view.is_shared,
        created_by=view.user_id,
        created_at=view.created_at,
    )


@router.delete("/views/{view_id}")
async def delete_saved_view(
    view_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.delete_view — Delete a saved view."""
    result = await db.execute(select(SavedView).where(SavedView.id == view_id))
    view = result.scalar_one_or_none()
    if not view:
        raise HTTPException(404, "View not found")
    await db.delete(view)
    await db.commit()
    return {"status": "deleted"}


# ── table.import_excel ─────────────────────────────────────────────────────


@router.post("/import", response_model=ImportDiffResponse)
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.import_excel — Upload Excel, build diff for review."""
    from openpyxl import load_workbook

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows_data = list(ws.iter_rows(values_only=True))
    if len(rows_data) < 2:
        raise HTTPException(400, "File has no data rows")

    headers = [str(h).strip() if h else "" for h in rows_data[0]]
    data_rows = rows_data[1:]

    # Try to find ID column (hidden)
    id_col = None
    for i, h in enumerate(headers):
        if not h:
            id_col = i
            break

    diff: list[ImportDiffRow] = []
    creates = 0
    updates = 0
    skips = 0
    errors = 0

    for idx, row in enumerate(data_rows):
        row_id = None
        if id_col is not None and len(row) > id_col and row[id_col]:
            try:
                row_id = uuid.UUID(str(row[id_col]))
            except (ValueError, AttributeError):
                pass

        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}

        if row_id:
            # Check if entity exists
            result = await db.execute(select(Invoice).where(Invoice.id == row_id))
            existing = result.scalar_one_or_none()
            if existing:
                changes = _detect_invoice_changes(existing, row_dict)
                if changes:
                    diff.append(ImportDiffRow(
                        row_index=idx, entity_id=row_id, action="update", changes=changes,
                    ))
                    updates += 1
                else:
                    diff.append(ImportDiffRow(row_index=idx, entity_id=row_id, action="skip"))
                    skips += 1
            else:
                diff.append(ImportDiffRow(row_index=idx, action="create", changes=row_dict))
                creates += 1
        else:
            diff.append(ImportDiffRow(row_index=idx, action="create", changes=row_dict))
            creates += 1

    import_id = uuid.uuid4()

    return ImportDiffResponse(
        import_id=import_id,
        file_name=file.filename or "upload.xlsx",
        total_rows=len(data_rows),
        creates=creates,
        updates=updates,
        skips=skips,
        errors=errors,
        diff=diff,
    )


def _detect_invoice_changes(invoice: Invoice, row_dict: dict) -> dict:
    changes = {}
    label_to_field = {
        "Номер счёта": "invoice_number",
        "Сумма": "total_amount",
        "НДС": "tax_amount",
        "Сумма без НДС": "subtotal",
        "Валюта": "currency",
    }
    for label, field in label_to_field.items():
        if label in row_dict and row_dict[label] is not None:
            current = getattr(invoice, field, None)
            new_val = row_dict[label]
            if field in ("total_amount", "tax_amount", "subtotal"):
                try:
                    new_val = float(new_val)
                except (ValueError, TypeError):
                    continue
            if str(current) != str(new_val):
                changes[field] = {"old": current, "new": new_val}
    return changes


# ── table.apply_diff ───────────────────────────────────────────────────────


class ApplyDiffRow(BaseModel):
    entity_id: uuid.UUID | None = None
    action: str  # update | create | skip
    changes: dict = {}


class ApplyDiffRequest(BaseModel):
    rows: list[ApplyDiffRow]


class ApplyDiffResult(BaseModel):
    applied: int
    skipped: int
    errors: list[str]


@router.post("/apply-diff", response_model=ApplyDiffResult)
async def apply_diff(
    payload: ApplyDiffRequest,
    db: AsyncSession = Depends(get_db),
    _user: UserInfo = Depends(require_role(UserRole.manager, UserRole.accountant)),
):
    """Skill: table.apply_diff — Apply import diff rows to the database."""
    applied = 0
    skipped = 0
    errors: list[str] = []

    numeric_fields = {"total_amount", "tax_amount", "subtotal", "unit_price"}

    for row in payload.rows:
        if row.action == "skip":
            skipped += 1
            continue

        try:
            if row.action == "update" and row.entity_id:
                result = await db.execute(
                    select(Invoice).where(Invoice.id == row.entity_id)
                )
                invoice = result.scalar_one_or_none()
                if not invoice:
                    errors.append(f"Invoice {row.entity_id} not found")
                    continue
                for field, change in row.changes.items():
                    new_val = change.get("new") if isinstance(change, dict) else change
                    if field in numeric_fields:
                        try:
                            new_val = float(new_val)
                        except (ValueError, TypeError):
                            continue
                    if hasattr(invoice, field):
                        setattr(invoice, field, new_val)
                applied += 1

            elif row.action == "create":
                # Create minimal Invoice record from changes
                inv = Invoice(**{
                    k: v for k, v in row.changes.items()
                    if hasattr(Invoice, k)
                })
                db.add(inv)
                applied += 1

        except Exception as e:
            errors.append(str(e))

    await db.commit()
    return ApplyDiffResult(applied=applied, skipped=skipped, errors=errors)


# ── Inline edit (table.inline_edit) ────────────────────────────────────────


class InlineEditRequest(BaseModel):
    entity_id: uuid.UUID
    field: str
    value: str | float | None


class InlineEditResponse(BaseModel):
    entity_id: uuid.UUID
    field: str
    old_value: str | float | None
    new_value: str | float | None


@router.post("/inline-edit", response_model=InlineEditResponse)
async def inline_edit(
    payload: InlineEditRequest,
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.inline_edit — Edit a single cell value."""
    from app.db.models import Invoice as InvoiceModel

    result = await db.execute(
        select(InvoiceModel).where(InvoiceModel.id == payload.entity_id)
    )
    invoice = result.scalar_one_or_none()
    if not invoice:
        raise HTTPException(404, "Entity not found")

    editable_fields = {
        "invoice_number", "currency", "total_amount", "tax_amount", "subtotal",
    }
    if payload.field not in editable_fields:
        raise HTTPException(400, f"Field '{payload.field}' is not editable")

    old_value = getattr(invoice, payload.field, None)

    # Type coercion
    new_value = payload.value
    if payload.field in ("total_amount", "tax_amount", "subtotal") and new_value is not None:
        new_value = float(new_value)

    setattr(invoice, payload.field, new_value)

    await log_action(
        db, action="table.inline_edit", entity_type="invoice",
        entity_id=invoice.id,
        details={"field": payload.field, "old": str(old_value), "new": str(new_value)},
    )
    await db.commit()

    return InlineEditResponse(
        entity_id=invoice.id, field=payload.field,
        old_value=old_value, new_value=new_value,
    )


# ── Batch actions ──────────────────────────────────────────────────────────


class BatchActionRequest(BaseModel):
    action: str  # approve, reject, delete
    entity_ids: list[uuid.UUID]
    reason: str | None = None


class BatchActionResponse(BaseModel):
    action: str
    total: int
    succeeded: int
    failed: int
    errors: list[str] = []


@router.post("/batch", response_model=BatchActionResponse)
async def batch_action(
    payload: BatchActionRequest,
    db: AsyncSession = Depends(get_db),
):
    """Skill: table.batch_action — Apply action to multiple invoices."""
    from app.audit.service import add_timeline_event as ate
    from app.db.models import Invoice as InvoiceModel
    from app.db.models import InvoiceStatus as InvoiceStatusModel

    succeeded = 0
    errors: list[str] = []

    for eid in payload.entity_ids:
        result = await db.execute(select(InvoiceModel).where(InvoiceModel.id == eid))
        inv = result.scalar_one_or_none()
        if not inv:
            errors.append(f"{eid}: not found")
            continue

        if payload.action == "approve":
            if inv.status not in (InvoiceStatusModel.needs_review, InvoiceStatusModel.draft):
                errors.append(f"{eid}: cannot approve (status={inv.status.value})")
                continue
            inv.status = InvoiceStatusModel.approved
            await ate(db, entity_type="invoice", entity_id=inv.id,
                      event_type="approved", summary="Batch approved", actor="user")
            succeeded += 1

        elif payload.action == "reject":
            if inv.status not in (InvoiceStatusModel.needs_review, InvoiceStatusModel.draft):
                errors.append(f"{eid}: cannot reject (status={inv.status.value})")
                continue
            inv.status = InvoiceStatusModel.rejected
            await ate(db, entity_type="invoice", entity_id=inv.id,
                      event_type="rejected",
                      summary=f"Batch rejected: {payload.reason or 'no reason'}",
                      actor="user")
            succeeded += 1

        else:
            errors.append(f"Unknown action: {payload.action}")
            break

    if succeeded > 0:
        await log_action(
            db, action=f"table.batch_{payload.action}", entity_type="invoice",
            entity_id=None,
            details={"count": succeeded, "ids": [str(e) for e in payload.entity_ids]},
        )
        await db.commit()

    return BatchActionResponse(
        action=payload.action,
        total=len(payload.entity_ids),
        succeeded=succeeded,
        failed=len(errors),
        errors=errors,
    )
