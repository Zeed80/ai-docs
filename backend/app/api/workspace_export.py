"""Professional exports for Workspace blocks."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.domain.workspace import get_workspace_block
from app.formatting import format_money, format_number, is_money_key, to_decimal

router = APIRouter()


class WorkspaceTableExportRequest(BaseModel):
    format: str = "xlsx"


@router.post("/blocks/{block_id}/export", response_class=StreamingResponse)
async def export_workspace_block(
    block_id: str,
    payload: WorkspaceTableExportRequest,
) -> StreamingResponse:
    """Export a Workspace table with professional formatting."""
    block = get_workspace_block(block_id)
    if not block:
        raise HTTPException(status_code=404, detail="Workspace block not found")
    if block.get("type") != "table":
        raise HTTPException(status_code=400, detail="Only table blocks can be exported")

    columns = [column for column in block.get("columns") or [] if isinstance(column, dict)]
    rows = [row for row in block.get("rows") or [] if isinstance(row, dict)]
    title = str(block.get("title") or "Рабочий стол")

    if payload.format == "csv":
        return _export_workspace_csv(columns, rows, title)
    return _export_workspace_xlsx(columns, rows, title)


def _export_workspace_xlsx(
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    title: str,
) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(title)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    numeric_alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    money_format = '# ##0,00'

    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(column.get("header") or column.get("key")))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, column in enumerate(columns, start=1):
            key = str(column.get("key") or "")
            col_type = str(column.get("type") or "text")
            value = _workspace_excel_value(row.get(key), key, col_type)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            link = _workspace_link(row.get(key))
            if link:
                cell.hyperlink = link
                cell.font = link_font
            if is_money_key(key):
                cell.number_format = money_format
                cell.alignment = numeric_alignment
            elif col_type == "number":
                cell.number_format = '# ##0,####'
                cell.alignment = numeric_alignment
        ws.row_dimensions[row_idx].height = _row_height(row, columns)

    for col_idx, column in enumerate(columns, start=1):
        key = str(column.get("key") or "")
        header = str(column.get("header") or key)
        width = max(len(header) + 3, 12)
        for row in rows:
            text = _workspace_display_value(row.get(key), key, str(column.get("type") or "text"))
            width = max(width, max((len(line) for line in text.splitlines()), default=0) + 3)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 70)

    if columns:
        last_col = get_column_letter(len(columns))
        last_row = max(len(rows) + 1, 1)
        table = Table(displayName="workspace_export", ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{_safe_filename(title)}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(filename)},
    )


def _export_workspace_csv(
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    title: str,
) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\n")
    writer.writerow([str(column.get("header") or column.get("key") or "") for column in columns])
    for row in rows:
        writer.writerow([
            _workspace_display_value(
                row.get(str(column.get("key") or "")),
                str(column.get("key") or ""),
                str(column.get("type") or "text"),
            )
            for column in columns
        ])
    output = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    filename = f"{_safe_filename(title)}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": _content_disposition(filename)},
    )


def _workspace_excel_value(value: Any, key: str, col_type: str) -> Any:
    if value is None:
        return None
    if isinstance(value, dict):
        label = str(value.get("label") or "").strip()
        href = str(value.get("href") or "").strip()
        return label if label and label.lower() not in {"скачать", "открыть"} else href
    if is_money_key(key) or col_type == "number":
        number = to_decimal(value)
        if number is not None:
            return float(number)
    return value


def _workspace_display_value(value: Any, key: str, col_type: str) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        label = str(value.get("label") or "").strip()
        href = str(value.get("href") or "").strip()
        if href and label:
            return f"{label}: {href}"
        return label or href
    if is_money_key(key):
        return format_money(value)
    if col_type == "number":
        return format_number(value)
    return str(value)


def _workspace_link(value: Any) -> str | None:
    if not isinstance(value, dict):
        return None
    href = value.get("href")
    return str(href) if isinstance(href, str) and href.strip() else None


def _row_height(row: dict[str, Any], columns: list[dict[str, Any]]) -> int:
    max_lines = 1
    for column in columns:
        key = str(column.get("key") or "")
        text = str(row.get(key) or "")
        max_lines = max(max_lines, text.count("\n") + 1)
    return min(max(18, max_lines * 16), 180)


def _safe_sheet_title(title: str) -> str:
    cleaned = "".join(ch for ch in title if ch not in r"[]:*?/\\").strip()
    return (cleaned or "Данные")[:31]


def _safe_filename(title: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in " _-" else "_" for ch in title).strip()
    return (cleaned or "workspace").replace(" ", "_")[:80]


def _content_disposition(filename: str) -> str:
    ascii_filename = "".join(
        ch if ch.isascii() and (ch.isalnum() or ch in "._-") else "_"
        for ch in filename
    ).strip("_")
    ascii_filename = ascii_filename or "workspace_export"
    return f"attachment; filename={ascii_filename}; filename*=UTF-8''{quote(filename)}"
