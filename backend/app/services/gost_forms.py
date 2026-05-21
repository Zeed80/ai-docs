"""ГОСТ ЕСТД form renderer for manufacturing tech processes.

Generates ГОСТ-compliant Excel workbooks:
  МК (Маршрутная карта) — ГОСТ 3.1118, форма 1 / 1а
  ОК (Операционная карта) — ГОСТ 3.1404, форма 2 / 2а
"""

from __future__ import annotations

import io
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# Row type codes per ГОСТ 3.1118
_ROW_M = "М"   # материал / заготовка
_ROW_O = "О"   # операция
_ROW_B = "Б"   # оснастка
_ROW_T = "Т"   # режущий инструмент / режимы

# GOST MK column widths (approximate, columns A-U based on ГОСТ 3.1118)
_MK_COL_WIDTHS: list[tuple[str, float]] = [
    ("A", 4),   # Тип строки
    ("B", 6),   # Цех
    ("C", 6),   # Уч.
    ("D", 6),   # РМ
    ("E", 8),   # Опер.
    ("F", 35),  # Наименование операции / материала
    ("G", 8),   # Оборудование (код)
    ("H", 16),  # Оборудование (модель)
    ("I", 8),   # СМ
    ("J", 6),   # Проф.
    ("K", 6),   # Разряд
    ("L", 6),   # Кол-во
    ("M", 8),   # Тпз, мин
    ("N", 8),   # Тшт, мин
    ("O", 8),   # Тшт-к, мин
]

_THIN = None  # filled after openpyxl import
_HEADER_FILL = None
_HEADER_FONT = None
_OP_FILL = None


def _init_styles():
    global _THIN, _HEADER_FILL, _HEADER_FONT, _OP_FILL
    if not _HAS_OPENPYXL:
        return
    thin = Side(style="thin")
    _THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
    _HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
    _HEADER_FONT = Font(bold=True, size=9)
    _OP_FILL = PatternFill("solid", fgColor="F2F2F2")


def _cell(ws, row: int, col: int, value: Any = "", bold: bool = False, fill=None, wrap: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    if _THIN:
        c.border = _THIN
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    if bold and _HEADER_FONT:
        c.font = Font(bold=True, size=9)
    else:
        c.font = Font(size=8)
    if fill:
        c.fill = fill
    return c


class GostFormRenderer:
    """Render ГОСТ ЕСТД forms МК and ОК as openpyxl workbooks."""

    def render_mk_xlsx(self, plan, operations: list) -> bytes:
        """Render МК (ГОСТ 3.1118) — route card workbook."""
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")
        _init_styles()
        wb = Workbook()
        ws = wb.active
        ws.title = "МК"

        # Column widths
        for col_letter, width in _MK_COL_WIDTHS:
            ws.column_dimensions[col_letter].width = width

        row = 1
        # ── Document header ───────────────────────────────────────────────
        ws.merge_cells(f"A{row}:O{row}")
        c = ws.cell(row=row, column=1, value="МАРШРУТНАЯ КАРТА")
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        headers = [
            ("ГОСТ 3.1118", "Форма 1"),
            (f"Изделие: {plan.product_name}", f"Обозначение: {plan.product_code or ''}"),
            (f"Тип ТП: {getattr(plan, 'tp_type', 'единичный')}", f"Версия/Литера: {plan.version}"),
            (f"Стандарт: {plan.standard_system}", f"Статус: {plan.status}"),
        ]
        for left, right in headers:
            ws.merge_cells(f"A{row}:H{row}")
            ws.cell(row=row, column=1, value=left).font = Font(size=8)
            ws.merge_cells(f"I{row}:O{row}")
            ws.cell(row=row, column=9, value=right).font = Font(size=8)
            row += 1

        # ── Material row (М) ──────────────────────────────────────────────
        material = plan.material or "—"
        blank_type = plan.blank_type or "—"
        _cell(ws, row, 1, _ROW_M, fill=_HEADER_FILL)
        ws.merge_cells(f"B{row}:H{row}")
        _cell(ws, row, 2, f"Материал: {material}. Заготовка: {blank_type}.")
        ws.merge_cells(f"I{row}:O{row}")
        _cell(ws, row, 9, f"Качество: {plan.quality_requirements or '—'}", wrap=True)
        ws.row_dimensions[row].height = 15
        row += 1

        # ── Column header row ─────────────────────────────────────────────
        mk_headers = ["Тип", "Цех", "Уч", "РМ", "Опер.", "Наименование операции",
                      "Оборуд.\n(код)", "Оборудование\n(модель)", "СМ", "Проф.", "Разр.",
                      "Кол.", "Тпз\nмин", "Тшт\nмин", "Тшт-к\nмин"]
        for col_idx, header in enumerate(mk_headers, start=1):
            _cell(ws, row, col_idx, header, bold=True, fill=_HEADER_FILL, wrap=True)
        ws.row_dimensions[row].height = 30
        row += 1

        total_tsht = 0.0
        total_tsht_k = 0.0

        for op in sorted(operations, key=lambda o: o.sequence_no):
            fill = _OP_FILL if op.operation_type == "quality_control" else None

            machine_code = ""
            machine_model = ""
            if op.machine_resource_id and hasattr(op, "machine_resource") and op.machine_resource:
                machine_code = op.machine_resource.code or ""
                machine_model = op.machine_resource.model or op.machine_resource.name or ""

            row_vals = [
                _ROW_O,
                "",   # цех
                "",   # участок
                "",   # рабочее место
                f"{op.sequence_no:03d}",
                op.name,
                machine_code,
                machine_model,
                "",   # СМ
                "",   # профессия
                "",   # разряд
                "",   # кол-во рабочих
                op.tpz_minutes or op.setup_minutes or "",
                op.tsht_minutes or op.labor_minutes or "",
                op.tsht_k_minutes or "",
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                _cell(ws, row, col_idx, val, fill=fill)
            ws.row_dimensions[row].height = 14
            row += 1

            if op.tsht_minutes:
                total_tsht += op.tsht_minutes
            if op.tsht_k_minutes:
                total_tsht_k += op.tsht_k_minutes

            # Tooling row (Б) if tooling_list exists
            if op.tooling_list:
                _cell(ws, row, 1, _ROW_B)
                tooling_str = "; ".join(
                    f"{t.get('name', '')} {t.get('code', '')}"
                    for t in (op.tooling_list or [])[:6]
                )
                ws.merge_cells(f"B{row}:O{row}")
                _cell(ws, row, 2, tooling_str, wrap=True)
                ws.row_dimensions[row].height = 12
                row += 1

            # Transition text row (Т) if cutting params
            if op.cutting_parameters:
                cp = op.cutting_parameters
                _cell(ws, row, 1, _ROW_T)
                params_str = (
                    f"Vc={cp.get('vc_m_min', '?')} м/мин  "
                    f"n={cp.get('n_rpm', '?')} об/мин  "
                    f"S={cp.get('feed_mm_min', '?')} мм/мин  "
                    f"t={cp.get('ap_mm', '?')} мм"
                )
                ws.merge_cells(f"B{row}:O{row}")
                _cell(ws, row, 2, params_str)
                ws.row_dimensions[row].height = 12
                row += 1

        # ── Totals row ────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:L{row}")
        _cell(ws, row, 1, "ИТОГО:", bold=True)
        _cell(ws, row, 13, "", bold=True)
        _cell(ws, row, 14, round(total_tsht, 2), bold=True)
        _cell(ws, row, 15, round(total_tsht_k, 2), bold=True)
        row += 1

        # ── Signature block ───────────────────────────────────────────────
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1, value="Разработал:").font = Font(size=8)
        ws.merge_cells(f"F{row}:J{row}")
        ws.cell(row=row, column=6, value="Проверил:").font = Font(size=8)
        ws.merge_cells(f"K{row}:O{row}")
        ws.cell(row=row, column=11, value="Утвердил:").font = Font(size=8)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def render_ok_xlsx(self, plan, operation) -> bytes:
        """Render one ОК (ГОСТ 3.1404 форма 2) for a single operation."""
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")
        _init_styles()
        wb = Workbook()
        ws = wb.active
        ws.title = f"ОК_{operation.sequence_no:03d}"

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 8

        row = 1
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value="ОПЕРАЦИОННАЯ КАРТА МЕХАНИЧЕСКОЙ ОБРАБОТКИ").font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1

        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value=f"ГОСТ 3.1404  Форма 2").font = Font(size=8, italic=True)
        row += 1

        info_rows = [
            (f"Изделие: {plan.product_name}", f"Обозначение: {plan.product_code or ''}"),
            (f"Материал: {plan.material or '—'}", f"Заготовка: {plan.blank_type or '—'}"),
            (f"Операция {operation.sequence_no:03d}: {operation.name}",
             f"Код: {operation.operation_code or operation.gost_operation_code or '—'}"),
        ]
        for left, right in info_rows:
            ws.merge_cells(f"A{row}:D{row}")
            ws.cell(row=row, column=1, value=left).font = Font(size=8)
            ws.merge_cells(f"E{row}:H{row}")
            ws.cell(row=row, column=5, value=right).font = Font(size=8)
            row += 1

        # Equipment
        machine_name = ""
        if hasattr(operation, "machine_resource") and operation.machine_resource:
            machine_name = operation.machine_resource.name or ""
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value=f"Оборудование: {machine_name}").font = Font(size=8)
        row += 2

        # Transitions header
        for col, header in enumerate(["№", "Код перехода", "Содержание перехода", "Т1 Режущий", "Т2 Вспом.", "Т3 Измер.", "То, мин", "Тв, мин"], start=1):
            c = ws.cell(row=row, column=col, value=header)
            c.font = Font(bold=True, size=8)
            c.fill = _HEADER_FILL
            c.border = _THIN
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28
        row += 1

        # Transition lines
        transitions = (operation.transition_text or "").strip().split("\n")
        to_per_step = round((operation.to_minutes or 0) / max(len(transitions), 1), 3)
        tv_per_step = round((operation.tv_minutes or 0) / max(len(transitions), 1), 3)

        for i, trans in enumerate(transitions, start=1):
            trans = trans.strip()
            if not trans:
                continue
            for col, val in enumerate([i, "", trans, "", "", "", to_per_step, tv_per_step], start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(size=8)
                c.border = _THIN
                c.alignment = Alignment(vertical="top", wrap_text=(col == 3))
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1
        # Norms summary
        norms = [
            ("То (машинное), мин", operation.to_minutes),
            ("Тв (вспомог.), мин", operation.tv_minutes),
            ("Тшт (штучное), мин", operation.tsht_minutes),
            ("Тшт-к (штучно-калькул.), мин", operation.tsht_k_minutes),
            ("Тпз (подг.-закл.), мин", operation.tpz_minutes),
        ]
        for label, val in norms:
            ws.cell(row=row, column=1, value=label).font = Font(size=8)
            ws.cell(row=row, column=2, value=val or "—").font = Font(size=8)
            row += 1

        # Cutting params
        if operation.cutting_parameters:
            row += 1
            ws.merge_cells(f"A{row}:H{row}")
            cp = operation.cutting_parameters
            params = (
                f"Режимы резания: Vc={cp.get('vc_m_min', '?')} м/мин, "
                f"n={cp.get('n_rpm', '?')} об/мин, "
                f"Sz={cp.get('fz_mm', '?')} мм/зуб, "
                f"t={cp.get('ap_mm', '?')} мм"
            )
            ws.cell(row=row, column=1, value=params).font = Font(size=8)
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1, value="Разработал:").font = Font(size=8)
        ws.merge_cells(f"D{row}:F{row}")
        ws.cell(row=row, column=4, value="Проверил:").font = Font(size=8)
        ws.merge_cells(f"G{row}:H{row}")
        ws.cell(row=row, column=7, value="Дата:").font = Font(size=8)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def render_full_package_xlsx(
        self, plan, operations: list, forms: list[str] | None = None
    ) -> bytes:
        """Render МК + ОК for all machining operations in one workbook."""
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed")
        _init_styles()

        if forms is None:
            forms = ["МК", "ОК"]

        wb = Workbook()
        # Remove default empty sheet
        default_ws = wb.active

        if "МК" in forms:
            mk_bytes = self.render_mk_xlsx(plan, operations)
            mk_wb = openpyxl.load_workbook(io.BytesIO(mk_bytes))
            mk_ws = mk_wb.active
            mk_ws.title = "МК"
            # Copy MK sheet into main workbook
            new_ws = wb.create_sheet("МК")
            for row in mk_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.alignment = cell.alignment.copy()
            for merge in mk_ws.merged_cells.ranges:
                new_ws.merge_cells(str(merge))
            for col_letter, col_dim in mk_ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = col_dim.width
            for row_dim_key, row_dim in mk_ws.row_dimensions.items():
                new_ws.row_dimensions[row_dim_key].height = row_dim.height

        if "ОК" in forms:
            machining_types = {"turning", "milling", "drilling", "grinding", "boring",
                               "reaming", "honing", "broaching"}
            for op in sorted(operations, key=lambda o: o.sequence_no):
                if op.operation_type not in machining_types:
                    continue
                ok_bytes = self.render_ok_xlsx(plan, op)
                ok_wb = openpyxl.load_workbook(io.BytesIO(ok_bytes))
                ok_source = ok_wb.active
                sheet_name = f"ОК_{op.sequence_no:03d}"
                new_ws = wb.create_sheet(sheet_name)
                for row in ok_source.iter_rows():
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = cell.font.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.alignment = cell.alignment.copy()
                for merge in ok_source.merged_cells.ranges:
                    new_ws.merge_cells(str(merge))
                for col_letter, col_dim in ok_source.column_dimensions.items():
                    new_ws.column_dimensions[col_letter].width = col_dim.width

        # Remove empty default sheet if other sheets were created
        if len(wb.sheetnames) > 1 and default_ws.title in wb.sheetnames:
            try:
                wb.remove(default_ws)
            except Exception:
                pass

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
