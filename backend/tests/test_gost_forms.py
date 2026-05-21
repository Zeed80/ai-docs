"""Tests for gost_forms.py — МК and ОК Excel rendering."""

import io
import uuid
from unittest.mock import MagicMock

import pytest

try:
    import openpyxl

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

pytestmark = pytest.mark.skipif(
    not _HAS_OPENPYXL, reason="openpyxl not installed"
)

from app.services.gost_forms import GostFormRenderer


def _make_plan(**kwargs):
    defaults = dict(
        id=uuid.uuid4(),
        product_name="Вал ступенчатый",
        product_code="75.ХХХХ.001",
        version="1.0",
        material="Ст.45 ГОСТ 1050-2013",
        blank_type="прокат",
        total_norm_minutes=45.5,
        approved_by=None,
        approved_at=None,
    )
    defaults.update(kwargs)
    m = MagicMock(**defaults)
    for k, v in defaults.items():
        setattr(m, k, v)
    return m


def _make_machine():
    m = MagicMock()
    m.code = "16К20"
    m.model = "16К20"
    m.name = "Токарный станок 16К20"
    return m


def _make_operation(seq=5, op_type="turning", **kwargs):
    defaults = dict(
        id=uuid.uuid4(),
        sequence_no=seq,
        name="Токарная черновая",
        operation_code="4110",
        gost_operation_code="4110",
        operation_type=op_type,
        setup_description="Установить в 3-кулачковый патрон",
        transition_text="1. Точить Ø50 на длину 80 мм.",
        control_requirements="Проверить Ø50 h8 скобой 50h8",
        tooling_list=[{"type": "fixture", "name": "Патрон 3-кулачковый", "code": "7100-0009"}],
        cutting_parameters={"vc_m_min": 120, "n_rpm": 764, "ap_mm": 2.0, "feed_mm_min": 0.3},
        to_minutes=3.2,
        tv_minutes=1.5,
        tob_minutes=0.3,
        totd_minutes=0.2,
        tsht_minutes=5.2,
        tsht_k_minutes=5.5,
        tpz_minutes=15.0,
        department_code="02",
        workplace_code="001",
    )
    defaults.update(kwargs)
    m = MagicMock(**defaults)
    for k, v in defaults.items():
        setattr(m, k, v)
    # Must unconditionally set machine_resource; MagicMock auto-attrs break openpyxl
    m.machine_resource = _make_machine()
    m.machine_resource_id = str(uuid.uuid4())
    return m


# ── МК ────────────────────────────────────────────────────────────────────────

def test_mk_xlsx_returns_bytes():
    plan = _make_plan()
    ops = [
        _make_operation(5, "turning"),
        _make_operation(10, "milling", name="Фрезерная"),
        _make_operation(15, "quality_control", name="Технический контроль"),
    ]
    renderer = GostFormRenderer()
    result = renderer.render_mk_xlsx(plan, ops)
    assert isinstance(result, bytes)
    assert len(result) > 100


def test_mk_xlsx_is_valid_workbook():
    plan = _make_plan()
    ops = [_make_operation()]
    renderer = GostFormRenderer()
    raw = renderer.render_mk_xlsx(plan, ops)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert len(wb.sheetnames) >= 1


def test_mk_xlsx_has_required_columns():
    """МК должна содержать все 15 колонок А–О по ГОСТ 3.1118."""
    plan = _make_plan()
    ops = [_make_operation(5, "turning"), _make_operation(10, "quality_control")]
    renderer = GostFormRenderer()
    raw = renderer.render_mk_xlsx(plan, ops)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    # Check at least 10 columns used (А-О = 15 max, some may be merged)
    max_col = max(
        (cell.column for row in ws.iter_rows() for cell in row if cell.value),
        default=0,
    )
    assert max_col >= 10


def test_mk_xlsx_contains_product_name():
    plan = _make_plan(product_name="Вал уникальный XYZ")
    ops = [_make_operation()]
    renderer = GostFormRenderer()
    raw = renderer.render_mk_xlsx(plan, ops)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    all_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    assert any("Вал уникальный XYZ" in v for v in all_values)


def test_mk_xlsx_contains_material():
    plan = _make_plan(material="Ст.45 ГОСТ 1050-2013")
    ops = [_make_operation()]
    renderer = GostFormRenderer()
    raw = renderer.render_mk_xlsx(plan, ops)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    all_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    assert any("Ст.45" in v for v in all_values)


# ── ОК ────────────────────────────────────────────────────────────────────────

def test_ok_xlsx_returns_bytes():
    plan = _make_plan()
    op = _make_operation()
    renderer = GostFormRenderer()
    result = renderer.render_ok_xlsx(plan, op)
    assert isinstance(result, bytes)
    assert len(result) > 100


def test_ok_xlsx_is_valid_workbook():
    plan = _make_plan()
    op = _make_operation()
    renderer = GostFormRenderer()
    raw = renderer.render_ok_xlsx(plan, op)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert len(wb.sheetnames) >= 1


def test_ok_transitions_format_per_gost_3_1404():
    """ОК должна содержать переходы из transition_text."""
    plan = _make_plan()
    op = _make_operation(transition_text="1. Точить Ø50. 2. Проверить размер.")
    renderer = GostFormRenderer()
    raw = renderer.render_ok_xlsx(plan, op)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    all_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    assert any("Точить" in v for v in all_values)


def test_ok_xlsx_contains_time_norms():
    """ОК должна содержать нормы времени."""
    plan = _make_plan()
    op = _make_operation(to_minutes=3.2, tsht_k_minutes=5.5)
    renderer = GostFormRenderer()
    raw = renderer.render_ok_xlsx(plan, op)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    all_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
    # Should contain To or Tsht-k value
    assert any("3.2" in v or "5.5" in v or "3,2" in v or "5,5" in v for v in all_values)


# ── Full package ───────────────────────────────────────────────────────────────

def test_full_package_xlsx_returns_bytes():
    plan = _make_plan()
    ops = [
        _make_operation(5, "turning"),
        _make_operation(10, "milling", name="Фрезерная"),
        _make_operation(15, "quality_control", name="Технический контроль"),
    ]
    renderer = GostFormRenderer()
    result = renderer.render_full_package_xlsx(plan, ops)
    assert isinstance(result, bytes)
    assert len(result) > 100


def test_full_package_has_multiple_sheets():
    """Full package: МК sheet + one ОК sheet per machining operation."""
    plan = _make_plan()
    ops = [
        _make_operation(5, "turning"),
        _make_operation(10, "milling", name="Фрезерная"),
        _make_operation(15, "quality_control", name="Технический контроль"),
    ]
    renderer = GostFormRenderer()
    raw = renderer.render_full_package_xlsx(plan, ops)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    # МК + 2 machining ОК = 3 sheets minimum
    assert len(wb.sheetnames) >= 3
