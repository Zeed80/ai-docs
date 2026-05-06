"""Tests for Workspace table export formatting."""

import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from app.domain.workspace import clear_workspace_blocks, upsert_workspace_block


@pytest.mark.asyncio
async def test_workspace_export_csv_uses_russian_decimal_separator(client: AsyncClient):
    clear_workspace_blocks()
    upsert_workspace_block(
        "test:export",
        {
            "id": "test:export",
            "type": "table",
            "title": "Экспорт счетов",
            "columns": [
                {"key": "invoice_number", "header": "Счет", "type": "text"},
                {"key": "total_amount", "header": "Сумма", "type": "number"},
            ],
            "rows": [{"invoice_number": "T-100", "total_amount": "10000.5"}],
        },
    )

    resp = await client.post(
        "/api/workspace/blocks/test%3Aexport/export",
        json={"format": "csv"},
    )

    assert resp.status_code == 200
    text = resp.content.decode("utf-8-sig")
    assert text.splitlines()[0] == "Счет;Сумма"
    assert "10 000,50" in text


@pytest.mark.asyncio
async def test_workspace_export_xlsx_wraps_and_styles_table(client: AsyncClient):
    clear_workspace_blocks()
    upsert_workspace_block(
        "test:styled-export",
        {
            "id": "test:styled-export",
            "type": "table",
            "title": "Сгруппированные товары",
            "columns": [
                {"key": "invoice_number", "header": "Счет", "type": "text"},
                {"key": "items", "header": "Товары", "type": "text"},
                {"key": "total_amount", "header": "Сумма", "type": "number"},
            ],
            "rows": [
                {
                    "invoice_number": "T-100",
                    "items": "Товар 1 - 2 шт\nТовар 2 - 3 шт",
                    "total_amount": "1500.25",
                }
            ],
        },
    )

    resp = await client.post(
        "/api/workspace/blocks/test%3Astyled-export/export",
        json={"format": "xlsx"},
    )

    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws["A1"].value == "Счет"
    assert ws["B2"].alignment.wrap_text is True
    assert ws["C2"].number_format == "# ##0,00"
    assert ws["A1"].border.bottom.style == "thin"
