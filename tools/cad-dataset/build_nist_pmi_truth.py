#!/usr/bin/env python3
"""Build semantic PMI truth from the official NIST CTC/FTC workbooks.

The workbook rows are authoritative semantic definitions.  PDFs, STEP files
and vector IR are linked as source assets, but their presence alone does not
prove a PMI-to-topology or PMI-to-drawing association.
"""

from __future__ import annotations

import argparse
import collections
import json
import pathlib
import re
import unicodedata
from typing import Any, Iterable

from openpyxl import load_workbook


SCHEMA_VERSION = "nist-pmi-truth/1.0"
SUITE_PATTERN = re.compile(r"NIST-(CTC|FTC)-PMI-Definitions\.xlsx", re.IGNORECASE)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return " ".join(text.split()).casefold()


def normalize_identifier(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return normalize_text(value).replace(" ", "")


def _relative_paths(paths: Iterable[pathlib.Path], source_root: pathlib.Path) -> list[str]:
    return sorted(str(path.relative_to(source_root)) for path in paths)


def _suite_from_workbook(path: pathlib.Path) -> str:
    match = SUITE_PATTERN.fullmatch(path.name)
    if not match:
        raise ValueError(f"unsupported NIST PMI workbook name: {path.name}")
    return match.group(1).lower()


def _case_prefix(suite: str, case_id: str) -> str:
    return f"nist_{suite}_{int(case_id):02d}_asme1_"


def _step_scope(path: pathlib.Path) -> str:
    header = path.read_text(encoding="utf-8", errors="ignore")[:4096].casefold()
    return "geometry_only" if "geometry only" in header else "unspecified"


def _drawing_page_candidates(
    drawing_assets: list[pathlib.Path], atc_id: str, source_root: pathlib.Path
) -> list[dict[str, Any]]:
    try:
        import fitz
    except ImportError:
        return []
    candidates = []
    for path in drawing_assets:
        try:
            document = fitz.open(path)
        except Exception:  # noqa: BLE001 - a malformed evidence PDF is simply unusable
            continue
        try:
            for page_index, page in enumerate(document):
                text = " ".join(page.get_text("text").split())
                match = re.search(
                    r"Includes Atomic Test Cases\s*-\s*([0-9, ]+)", text, re.IGNORECASE
                )
                page_atc_ids = {
                    normalize_identifier(value)
                    for value in (match.group(1).split(",") if match else [])
                    if normalize_identifier(value)
                }
                if atc_id in page_atc_ids:
                    candidates.append(
                        {
                            "asset": str(path.relative_to(source_root)),
                            "page": page_index + 1,
                            "page_bbox": [0.0, 0.0, float(page.rect.width), float(page.rect.height)],
                            "basis": "official_atomic_test_case_page_membership",
                        }
                    )
        finally:
            document.close()
    return candidates


def parse_workbook(
    workbook_path: pathlib.Path,
    source_root: pathlib.Path,
    ir_root: pathlib.Path | None = None,
) -> list[dict[str, Any]]:
    suite = _suite_from_workbook(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_number <= 2:
            continue
        primary = normalize_identifier(row[0] if len(row) > 0 else None)
        atc = normalize_identifier(row[1] if len(row) > 1 else None)
        category = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        description = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        specification = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        if not primary or not atc or not category or not description:
            continue

        case_prefix = _case_prefix(suite, primary)
        case_assets = [
            path
            for path in workbook_path.parent.iterdir()
            if path.is_file() and path.name.casefold().startswith(case_prefix)
        ]
        step_assets = [path for path in case_assets if path.suffix.casefold() in {".stp", ".step"}]
        drawing_assets = [
            path
            for path in case_assets
            if path.suffix.casefold() == ".pdf"
            and "elem_ids" not in path.name.casefold()
            and "_fsi" not in path.name.casefold()
        ]
        element_id_assets = [
            path for path in case_assets if "elem_ids" in path.name.casefold()
        ]
        ir_assets: list[pathlib.Path] = []
        if ir_root and ir_root.exists():
            ir_assets = sorted(ir_root.glob(f"{case_prefix}*.json"))

        semantic_id = f"nist:{suite}:{int(primary):02d}:atc:{atc}"
        measurand = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        comments = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
        standards = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
        # FTC has seven columns and stores standards in column G.
        if suite == "ftc":
            standards = comments
            comments = measurand
            measurand = ""

        records.append(
            {
                "schema_version": SCHEMA_VERSION,
                "semantic_id": semantic_id,
                "suite": suite,
                "primary_case_id": str(int(primary)),
                "atc_id": atc,
                "category": category,
                "description": description,
                "specification": specification,
                "measurand": measurand,
                "comments": comments,
                "standards_mapping": standards,
                "normalized": {
                    "category": normalize_text(category),
                    "description": normalize_text(description),
                    "specification": normalize_text(specification),
                },
                "evidence": {
                    "authority": "NIST MBE PMI Definitions workbook",
                    "workbook": str(workbook_path.relative_to(source_root)),
                    "worksheet": worksheet.title,
                    "row": row_number,
                    "step_assets": _relative_paths(step_assets, source_root),
                    "step_scopes": {
                        str(path.relative_to(source_root)): _step_scope(path)
                        for path in step_assets
                    },
                    "drawing_assets": _relative_paths(drawing_assets, source_root),
                    "drawing_page_candidates": _drawing_page_candidates(
                        drawing_assets, atc, source_root
                    ),
                    "element_id_assets": _relative_paths(element_id_assets, source_root),
                    "ir_pages": (
                        [str(path.relative_to(ir_root.parent)) for path in ir_assets]
                        if ir_root
                        else []
                    ),
                },
                "assurance": {
                    "semantic_status": "source_defined",
                    "geometry_linked": False,
                    "drawing_located": False,
                    "unresolved": [
                        "pmi_to_topology_association_not_available_in_geometry_only_step",
                        "pmi_to_exact_drawing_region_association",
                    ],
                },
            }
        )
    workbook.close()
    return records


def validate_records(records: list[dict[str, Any]]) -> None:
    if not records:
        raise ValueError("PMI truth is empty")
    required = {
        "schema_version",
        "semantic_id",
        "suite",
        "primary_case_id",
        "atc_id",
        "category",
        "description",
        "specification",
        "evidence",
        "assurance",
    }
    seen: set[str] = set()
    for index, record in enumerate(records, start=1):
        missing = sorted(field for field in required if field not in record)
        if missing:
            raise ValueError(f"PMI truth record {index} missing fields: {', '.join(missing)}")
        semantic_id = record["semantic_id"]
        if semantic_id in seen:
            raise ValueError(f"duplicate PMI semantic_id: {semantic_id}")
        seen.add(semantic_id)
        if record["schema_version"] != SCHEMA_VERSION:
            raise ValueError(f"unsupported PMI truth schema: {record['schema_version']}")
        if record["assurance"].get("geometry_linked") and not record["evidence"].get("topology_targets"):
            raise ValueError(f"{semantic_id} claims geometry link without topology targets")
        if record["assurance"].get("drawing_located") and not record["evidence"].get("drawing_regions"):
            raise ValueError(f"{semantic_id} claims drawing location without regions")


def build_truth(source_root: pathlib.Path, ir_root: pathlib.Path | None = None) -> list[dict[str, Any]]:
    workbooks = sorted(source_root.rglob("NIST-*-PMI-Definitions.xlsx"))
    if not workbooks:
        raise FileNotFoundError(f"no NIST PMI definition workbooks under {source_root}")
    records = [
        record
        for workbook in workbooks
        for record in parse_workbook(workbook, source_root=source_root, ir_root=ir_root)
    ]
    records.sort(key=lambda item: (item["suite"], int(item["primary_case_id"]), item["atc_id"]))
    validate_records(records)
    return records


def summarize(records: list[dict[str, Any]]) -> dict[str, Any]:
    by_suite = collections.Counter(record["suite"] for record in records)
    by_category = collections.Counter(record["category"] for record in records)
    by_case = collections.Counter(
        f"{record['suite']}:{int(record['primary_case_id']):02d}" for record in records
    )
    return {
        "schema_version": SCHEMA_VERSION,
        "records": len(records),
        "suites": dict(sorted(by_suite.items())),
        "cases": len(by_case),
        "records_by_case": dict(sorted(by_case.items())),
        "records_by_category": dict(sorted(by_category.items())),
        "geometry_linked": sum(record["assurance"]["geometry_linked"] for record in records),
        "drawing_located": sum(record["assurance"]["drawing_located"] for record in records),
        "drawing_page_candidates": sum(
            bool(record["evidence"].get("drawing_page_candidates")) for record in records
        ),
        "drawing_page_candidate_links": sum(
            len(record["evidence"].get("drawing_page_candidates", [])) for record in records
        ),
        "geometry_only_step_records": sum(
            "geometry_only" in record["evidence"].get("step_scopes", {}).values()
            for record in records
        ),
        "promotion_ready": all(
            record["assurance"]["geometry_linked"] and record["assurance"]["drawing_located"]
            for record in records
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-root", type=pathlib.Path, required=True)
    parser.add_argument("--ir-root", type=pathlib.Path)
    parser.add_argument("--output", type=pathlib.Path, required=True)
    parser.add_argument("--summary", type=pathlib.Path, required=True)
    args = parser.parse_args()

    records = build_truth(args.source_root, args.ir_root)
    report = summarize(records)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        "".join(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n" for record in records),
        encoding="utf-8",
    )
    args.summary.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
